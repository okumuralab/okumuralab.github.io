import pathlib, re
from openpyxl import load_workbook
import numpy as np
import pandas as pd

# wget --timestamping --recursive --no-parent --wait=1 https://www.nier.go.jp/21chousakekkahoukoku/index.html

path = pathlib.Path("www.nier.go.jp/21chousakekkahoukoku/factsheet")

# 小学校
names = sorted([p for p in path.glob("**/*.xlsx") if re.search(r"/\d\dp_21r\.xlsx$", str(p))])

def f(n):
    wb = load_workbook(n, read_only=True)
    ws = wb.worksheets[0]['AE8:AE22']
    n = len(ws)
    x = np.array([ws[i][0].value for i in range(n)])
    koku = (x @ np.arange(n)[::-1]) / np.sum(x) * (100 / (n-1))
    ws = wb.worksheets[1]['AE8:AE24']
    n = len(ws)
    x = np.array([ws[i][0].value for i in range(n)])
    san = (x @ np.arange(n)[::-1]) / np.sum(x) * (100 / (n-1))
    return [koku, san]

sho = np.array([f(n) for n in names])

# 中学校
names = sorted([p for p in path.glob("**/*.xlsx") if re.search(r"/\d\dm_21r\.xlsx$", str(p))])

def f(n):
    wb = load_workbook(n, read_only=True)
    ws = wb.worksheets[0]['AE8:AE22']
    n = len(ws)
    x = np.array([ws[i][0].value for i in range(n)])
    koku = (x @ np.arange(n)[::-1]) / np.sum(x) * (100 / (n-1))
    ws = wb.worksheets[1]['AE8:AE24']
    n = len(ws)
    x = np.array([ws[i][0].value for i in range(n)])
    suu = (x @ np.arange(n)[::-1]) / np.sum(x) * (100 / (n-1))
    # ws = wb.worksheets[2]['AE8:AE29']
    # n = len(ws)
    # x = np.array([ws[i][0].value for i in range(n)])
    # ei = (x @ np.arange(n)[::-1]) / np.sum(x) * (100 / (n-1))
    # return [koku, suu, ei]
    return [koku, suu]

chu = np.array([f(n) for n in names])

shochu = np.hstack((sho, chu))

pd.DataFrame(shochu).to_csv("atest2021.csv", index=False,
                            encoding="utf_8_sig", line_terminator="\r\n",
                            header=["小国","小算","中国","中数"])
